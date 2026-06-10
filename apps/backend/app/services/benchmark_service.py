"""Benchmark service: CSV parsing, LLM matching, recommendation chat."""
from __future__ import annotations

import csv
import io
import json
import logging
import statistics
import uuid
from datetime import datetime, timezone
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.agent_runtime.providers import get_provider
from app.models.benchmark import (
    BenchmarkDataset,
    BenchmarkDatasetRow,
    BenchmarkMatch,
    BenchmarkRecommendation,
    JobListingBenchmarkRun,
)
from app.models.case import NegotiationCase

logger = logging.getLogger(__name__)

# Canonical field names used for column mapping
CANONICAL_FIELDS = [
    "title",
    "level",
    "department",
    "location",
    "country",
    "currency",
    "base_salary",
    "total_compensation",
    "bonus",
    "equity",
    "effective_date",
]

# Heuristic header → canonical field mapping for auto-detection
_HEADER_HINTS: dict[str, str] = {
    "job title": "title",
    "jobtitle": "title",
    "position": "title",
    "role": "title",
    "job_title": "title",
    "title": "title",
    "level": "level",
    "band": "level",
    "grade": "level",
    "seniority": "level",
    "job_level": "level",
    "department": "department",
    "dept": "department",
    "team": "department",
    "function": "department",
    "location": "location",
    "office": "location",
    "city": "location",
    "work location": "location",
    "work_location": "location",
    "country": "country",
    "country_code": "country",
    "currency": "currency",
    "pay_currency": "currency",
    "base salary": "base_salary",
    "base_salary": "base_salary",
    "base": "base_salary",
    "salary": "base_salary",
    "annual salary": "base_salary",
    "annual_salary": "base_salary",
    "total comp": "total_compensation",
    "total_comp": "total_compensation",
    "total compensation": "total_compensation",
    "total_compensation": "total_compensation",
    "total package": "total_compensation",
    "bonus": "bonus",
    "bonus_amount": "bonus",
    "annual bonus": "bonus",
    "equity": "equity",
    "equity_value": "equity",
    "equity value": "equity",
    "rsu": "equity",
    "effective date": "effective_date",
    "effective_date": "effective_date",
    "date": "effective_date",
    "pay_date": "effective_date",
}


def _infer_mapping(headers: list[str]) -> dict[str, str]:
    """Auto-detect canonical → CSV header mapping from header names."""
    mapping: dict[str, str] = {}
    for h in headers:
        canonical = _HEADER_HINTS.get(h.strip().lower())
        if canonical and canonical not in mapping:
            mapping[canonical] = h
    return mapping


def _normalize_headers(headers: list[str]) -> list[str]:
    normalized: list[str] = []
    seen: dict[str, int] = {}
    for idx, header in enumerate(headers, start=1):
        candidate = header.strip() if header.strip() else f"column_{idx}"
        seen_count = seen.get(candidate, 0)
        seen[candidate] = seen_count + 1
        normalized.append(candidate if seen_count == 0 else f"{candidate}_{seen_count + 1}")
    return normalized


def _parse_csv_rows(content: bytes) -> tuple[list[str], list[dict[str, str]]]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    headers = list(reader.fieldnames or [])
    rows: list[dict[str, str]] = []
    for raw_row in reader:
        rows.append({
            key: "" if value is None else str(value).strip()
            for key, value in dict(raw_row).items()
        })
    return headers, rows


def _parse_xlsx_rows(content: bytes) -> tuple[list[str], list[dict[str, str]]]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    header_row = next(iterator, None)
    if header_row is None:
        workbook.close()
        return [], []

    headers = _normalize_headers(["" if cell is None else str(cell).strip() for cell in header_row])
    rows: list[dict[str, str]] = []
    for row in iterator:
        row_dict: dict[str, str] = {}
        for idx, header in enumerate(headers):
            value = row[idx] if idx < len(row) else ""
            row_dict[header] = "" if value is None else str(value).strip()
        rows.append(row_dict)

    workbook.close()
    return headers, rows


def _parse_dataset_rows(content: bytes, filename: str) -> tuple[list[str], list[dict[str, str]]]:
    extension = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if extension == "csv":
        return _parse_csv_rows(content)
    if extension == "xlsx":
        return _parse_xlsx_rows(content)
    raise ValueError("Unsupported file type. Only .csv and .xlsx are accepted.")


def _safe_float(value: Any) -> float | None:
    if value is None or str(value).strip() == "":
        return None
    cleaned = str(value).replace(",", "").replace("$", "").replace("£", "").replace("€", "").strip()
    try:
        return float(cleaned)
    except (ValueError, TypeError):
        return None


def _normalize_currency(raw: str | None) -> str | None:
    if not raw:
        return None
    upper = str(raw).strip().upper()
    known = {"USD", "GBP", "EUR", "CAD", "AUD", "SGD", "INR", "JPY", "CHF", "NZD"}
    return upper if upper in known else None


def _apply_mapping_to_row(row: dict[str, Any], mapping: dict[str, str]) -> dict[str, Any]:
    """Extract canonical fields from a raw row using the column mapping."""
    def get(canonical: str) -> str | None:
        col = mapping.get(canonical)
        if not col:
            return None
        value = row.get(col)
        if value is None:
            return None
        text = str(value).strip()
        return text if text else None

    return {
        "normalized_title": get("title"),
        "normalized_level": get("level"),
        "department": get("department"),
        "location": get("location"),
        "country": get("country"),
        "currency": _normalize_currency(get("currency")),
        "base_salary": _safe_float(get("base_salary")),
        "total_compensation": _safe_float(get("total_compensation")),
        "bonus": _safe_float(get("bonus")),
        "equity": _safe_float(get("equity")),
        "effective_date": get("effective_date"),
    }


def _listing_context(listing: NegotiationCase) -> dict[str, Any]:
    """Extract key fields from a NegotiationCase for use in LLM prompts."""
    company_public: dict = {}
    company_confidential: dict = {}
    candidate_public: dict = {}
    for p in listing.parties:
        if p.party_type == "company":
            company_public = p.public_payload or {}
            company_confidential = p.confidential_payload or {}
        elif p.party_type == "candidate":
            candidate_public = p.public_payload or {}
    return {
        "title": listing.title,
        "description": listing.description or "",
        "job_title": company_public.get("job_title") or listing.title,
        "job_description": company_public.get("job_description", ""),
        "responsibilities": company_public.get("responsibilities", []),
        "location": company_public.get("location", ""),
        "work_arrangement": company_public.get("work_arrangement", ""),
        "category": company_public.get("category", ""),
        "currency": listing.currency or "USD",
        "jurisdiction": listing.jurisdiction or "US",
        "budget_floor": company_confidential.get("budget_floor"),
        "budget_target": company_confidential.get("budget_target"),
        "budget_ceiling": company_confidential.get("budget_ceiling"),
        "candidate_skills": candidate_public.get("strengths", []),
        "desired_base": (candidate_public.get("desired_compensation") or {}).get("base_salary_target"),
    }


def _extract_json(text: str) -> Any:
    """Extract JSON from potentially prose-wrapped LLM output."""
    text = text.strip()
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass
    # Try extracting from within ```json ... ``` blocks
    for delim in ["```json", "```"]:
        if delim in text:
            start = text.find(delim) + len(delim)
            end = text.rfind("```")
            if end > start:
                try:
                    return json.loads(text[start:end].strip())
                except json.JSONDecodeError:
                    pass
    # Try to find the outermost { } or [ ]
    for open_ch, close_ch in [("{", "}"), ("[", "]")]:
        s = text.find(open_ch)
        e = text.rfind(close_ch)
        if s != -1 and e != -1 and e > s:
            try:
                return json.loads(text[s : e + 1])
            except json.JSONDecodeError:
                pass
    return None


class BenchmarkService:
    def __init__(self, db: Session) -> None:
        self.db = db

    # ── Dataset management ────────────────────────────────────────────────────

    def list_datasets(self, tenant_id: uuid.UUID) -> list[BenchmarkDataset]:
        return list(
            self.db.scalars(
                select(BenchmarkDataset)
                .where(BenchmarkDataset.tenant_id == tenant_id)
                .where(BenchmarkDataset.is_active.is_(True))
                .order_by(BenchmarkDataset.created_at.desc())
            )
        )

    def get_dataset(self, dataset_id: uuid.UUID, tenant_id: uuid.UUID) -> BenchmarkDataset | None:
        return self.db.scalar(
            select(BenchmarkDataset)
            .where(BenchmarkDataset.id == dataset_id)
            .where(BenchmarkDataset.tenant_id == tenant_id)
        )

    def parse_and_store_dataset(
        self,
        content: bytes,
        filename: str,
        source_type: str,
        dataset_name: str,
        uploaded_by_id: uuid.UUID,
        tenant_id: uuid.UUID,
    ) -> BenchmarkDataset:
        """Parse a CSV/XLSX file and persist its rows to the database."""
        headers, parsed_rows = _parse_dataset_rows(content, filename)
        inferred_mapping = _infer_mapping(headers)
        has_mapping = bool(inferred_mapping)

        dataset = BenchmarkDataset(
            source_type=source_type,
            dataset_name=dataset_name,
            original_filename=filename,
            uploaded_by=uploaded_by_id,
            column_mapping_json=inferred_mapping if has_mapping else None,
            status="mapped" if has_mapping else "uploaded",
            is_global=True,
            is_active=True,
            tenant_id=tenant_id,
        )
        self.db.add(dataset)
        self.db.flush()

        row_count = 0
        for raw_row in parsed_rows:
            raw_dict = {
                key: "" if value is None else str(value).strip()
                for key, value in raw_row.items()
            }
            if not any(value.strip() for value in raw_dict.values()):
                continue  # skip blank rows
            normalized = _apply_mapping_to_row(raw_dict, inferred_mapping) if has_mapping else {
                "normalized_title": None,
                "normalized_level": None,
                "department": None,
                "location": None,
                "country": None,
                "currency": None,
                "base_salary": None,
                "total_compensation": None,
                "bonus": None,
                "equity": None,
                "effective_date": None,
            }
            self.db.add(
                BenchmarkDatasetRow(
                    dataset_id=dataset.id,
                    raw_row_json=raw_dict,
                    **normalized,
                )
            )
            row_count += 1

        dataset.row_count = row_count
        self.db.flush()
        return dataset

    def update_column_mapping(
        self,
        dataset_id: uuid.UUID,
        mapping: dict[str, str],
        tenant_id: uuid.UUID,
    ) -> BenchmarkDataset:
        """Save a column mapping and re-normalise existing rows."""
        dataset = self.get_dataset(dataset_id, tenant_id)
        if dataset is None:
            raise ValueError("Dataset not found")

        dataset.column_mapping_json = mapping
        dataset.status = "mapped"
        self.db.flush()

        # Re-normalise all rows with the new mapping
        rows = list(
            self.db.scalars(select(BenchmarkDatasetRow).where(BenchmarkDatasetRow.dataset_id == dataset_id))
        )
        for row in rows:
            normalized = _apply_mapping_to_row(row.raw_row_json, mapping)
            for field, value in normalized.items():
                setattr(row, field, value)

        self.db.flush()
        return dataset

    def get_dataset_rows(
        self, dataset_id: uuid.UUID, tenant_id: uuid.UUID, limit: int = 100, offset: int = 0
    ) -> tuple[int, list[BenchmarkDatasetRow]]:
        dataset = self.get_dataset(dataset_id, tenant_id)
        if dataset is None:
            raise ValueError("Dataset not found")
        total = dataset.row_count
        rows = list(
            self.db.scalars(
                select(BenchmarkDatasetRow)
                .where(BenchmarkDatasetRow.dataset_id == dataset_id)
                .order_by(BenchmarkDatasetRow.id)
                .offset(offset)
                .limit(limit)
            )
        )
        return total, rows

    def deactivate_dataset(self, dataset_id: uuid.UUID, tenant_id: uuid.UUID) -> BenchmarkDataset:
        dataset = self.get_dataset(dataset_id, tenant_id)
        if dataset is None:
            raise ValueError("Dataset not found")
        dataset.is_active = False
        self.db.flush()
        return dataset

    # ── Benchmark runs ────────────────────────────────────────────────────────

    def list_runs(self, job_listing_id: uuid.UUID, tenant_id: uuid.UUID) -> list[JobListingBenchmarkRun]:
        return list(
            self.db.scalars(
                select(JobListingBenchmarkRun)
                .where(JobListingBenchmarkRun.job_listing_id == job_listing_id)
                .where(JobListingBenchmarkRun.tenant_id == tenant_id)
                .order_by(JobListingBenchmarkRun.created_at.desc())
            )
        )

    def get_run(self, run_id: uuid.UUID, tenant_id: uuid.UUID) -> JobListingBenchmarkRun | None:
        return self.db.scalar(
            select(JobListingBenchmarkRun)
            .where(JobListingBenchmarkRun.id == run_id)
            .where(JobListingBenchmarkRun.tenant_id == tenant_id)
        )

    async def run_internal_benchmark(
        self,
        job_listing_id: uuid.UUID,
        dataset_ids: list[uuid.UUID],
        minimum_cohort: int,
        suppress_exact: bool,
        created_by_id: uuid.UUID,
        tenant_id: uuid.UUID,
    ) -> JobListingBenchmarkRun:
        listing = self.db.scalar(select(NegotiationCase).where(NegotiationCase.id == job_listing_id))
        if listing is None:
            raise ValueError("Job listing not found")

        run = JobListingBenchmarkRun(
            job_listing_id=job_listing_id,
            run_type="internal",
            status="running",
            created_by=created_by_id,
            input_params_json={
                "dataset_ids": [str(d) for d in dataset_ids],
                "minimum_cohort": minimum_cohort,
                "suppress_exact": suppress_exact,
            },
            tenant_id=tenant_id,
        )
        self.db.add(run)
        self.db.flush()

        try:
            listing_ctx = _listing_context(listing)
            all_rows: list[BenchmarkDatasetRow] = []
            for ds_id in dataset_ids:
                rows = list(
                    self.db.scalars(
                        select(BenchmarkDatasetRow).where(BenchmarkDatasetRow.dataset_id == ds_id)
                    )
                )
                all_rows.extend(rows)

            if not all_rows:
                run.status = "completed"
                run.completed_at = datetime.now(timezone.utc)
                run.result_summary_json = {"warning": "No dataset rows found. Upload and map a dataset first."}
                run.confidence_score = 0.0
                self.db.flush()
                return run

            # Build compact row summaries for LLM (max 200 rows)
            row_summaries = [
                {
                    "idx": i,
                    "title": r.normalized_title or "",
                    "level": r.normalized_level or "",
                    "dept": r.department or "",
                    "loc": r.location or "",
                    "currency": r.currency or listing_ctx["currency"],
                    "base": r.base_salary,
                    "total": r.total_compensation,
                }
                for i, r in enumerate(all_rows[:200])
            ]

            provider = get_provider()
            system_prompt = (
                "You are a compensation analyst. You will receive a job listing and a list of internal "
                "compensation records. Identify which records are the closest matches to the listing based on "
                "title, seniority, department, location, and responsibilities.\n"
                "Return ONLY valid JSON with this structure:\n"
                '{"matches": [{"idx": int, "confidence": float, "rationale": str}], '
                '"cohort_explanation": str}\n'
                "Include up to 30 best matches. Confidence is 0.0-1.0. Be strict about relevance."
            )
            user_content = (
                f"Job listing:\n{json.dumps(listing_ctx, indent=2)}\n\n"
                f"Internal records ({len(row_summaries)} rows):\n{json.dumps(row_summaries, indent=2)}"
            )
            result = await provider.generate(
                system_prompt=system_prompt,
                messages=[{"role": "user", "content": user_content}],
                temperature=0.1,
            )
            parsed = _extract_json(result.get("content", ""))
            llm_matches: list[dict] = (parsed.get("matches") or []) if isinstance(parsed, dict) else []
            cohort_explanation: str = (parsed.get("cohort_explanation") or "") if isinstance(parsed, dict) else ""

            matched_rows: list[BenchmarkDatasetRow] = []
            for m in llm_matches:
                idx = m.get("idx")
                if isinstance(idx, int) and 0 <= idx < len(all_rows):
                    matched_rows.append(all_rows[idx])
                    self.db.add(
                        BenchmarkMatch(
                            benchmark_run_id=run.id,
                            dataset_id=all_rows[idx].dataset_id,
                            source_type="internal",
                            matched_title=all_rows[idx].normalized_title,
                            matched_level=all_rows[idx].normalized_level,
                            matched_location=all_rows[idx].location,
                            base_salary=all_rows[idx].base_salary,
                            total_compensation=all_rows[idx].total_compensation,
                            currency=all_rows[idx].currency or listing_ctx["currency"],
                            confidence_score=float(m.get("confidence") or 0),
                            match_rationale=str(m.get("rationale") or ""),
                        )
                    )

            # Compute percentile summary
            base_salaries = [r.base_salary for r in matched_rows if r.base_salary is not None]
            total_comps = [r.total_compensation for r in matched_rows if r.total_compensation is not None]
            cohort_size = len(base_salaries)
            suppress = suppress_exact and cohort_size < minimum_cohort

            def _pct(values: list[float], p: float) -> float | None:
                if not values:
                    return None
                sorted_v = sorted(values)
                idx_f = (p / 100) * (len(sorted_v) - 1)
                lo = int(idx_f)
                hi = min(lo + 1, len(sorted_v) - 1)
                return sorted_v[lo] + (sorted_v[hi] - sorted_v[lo]) * (idx_f - lo)

            summary: dict[str, Any] = {
                "cohort_size": cohort_size,
                "minimum_cohort": minimum_cohort,
                "suppressed": suppress,
                "cohort_explanation": cohort_explanation,
            }
            if not suppress and base_salaries:
                summary["median_base"] = statistics.median(base_salaries)
                summary["p25_base"] = _pct(base_salaries, 25)
                summary["p75_base"] = _pct(base_salaries, 75)
                summary["min_base"] = min(base_salaries) if cohort_size >= 10 else None
                summary["max_base"] = max(base_salaries) if cohort_size >= 10 else None
                if total_comps:
                    summary["median_total_comp"] = statistics.median(total_comps)
                    summary["p75_total_comp"] = _pct(total_comps, 75)
            elif suppress:
                summary["guidance"] = (
                    f"Cohort too small ({cohort_size} < {minimum_cohort}). "
                    "Exact salary values are suppressed to protect employee confidentiality. "
                    "Add more data or reduce the minimum cohort threshold."
                )

            overall_confidence = (
                statistics.mean([m.get("confidence", 0) for m in llm_matches]) if llm_matches else 0.0
            )
            run.status = "completed"
            run.completed_at = datetime.now(timezone.utc)
            run.result_summary_json = summary
            run.confidence_score = overall_confidence
            self.db.flush()

        except Exception as exc:
            logger.exception("Internal benchmark run failed: %s", exc)
            run.status = "failed"
            run.result_summary_json = {"error": str(exc)}
            self.db.flush()

        return run

    async def run_external_benchmark(
        self,
        job_listing_id: uuid.UUID,
        sources: list[str],
        dataset_ids: list[uuid.UUID],
        search_params: dict[str, Any],
        created_by_id: uuid.UUID,
        tenant_id: uuid.UUID,
    ) -> JobListingBenchmarkRun:
        listing = self.db.scalar(select(NegotiationCase).where(NegotiationCase.id == job_listing_id))
        if listing is None:
            raise ValueError("Job listing not found")

        run = JobListingBenchmarkRun(
            job_listing_id=job_listing_id,
            run_type="external",
            status="running",
            created_by=created_by_id,
            input_params_json={"sources": sources, "dataset_ids": [str(d) for d in dataset_ids], "search_params": search_params},
            tenant_id=tenant_id,
        )
        self.db.add(run)
        self.db.flush()

        try:
            listing_ctx = _listing_context(listing)
            provider = get_provider()
            all_evidence: list[dict] = []

            # ── Web search (LLM-based) ─────────────────────────────────────
            if "web_search" in sources:
                system_prompt = (
                    "You are a compensation research specialist. Using your knowledge of salary data from "
                    "public sources (Glassdoor, Levels.fyi, Bureau of Labor Statistics, LinkedIn Salary, "
                    "Payscale, Radford, and similar), research realistic salary ranges for the given job.\n"
                    "Return ONLY valid JSON:\n"
                    '{"datapoints": [{"source_name": str, "matched_title": str, "geography": str, '
                    '"currency": str, "p50_base": float|null, "p75_base": float|null, '
                    '"total_comp_p50": float|null, "confidence": float, '
                    '"citation_note": str, "notes": str}]}'
                )
                user_content = (
                    f"Research external salary benchmarks for this role:\n{json.dumps(listing_ctx, indent=2)}"
                )
                result = await provider.generate(
                    system_prompt=system_prompt,
                    messages=[{"role": "user", "content": user_content}],
                    temperature=0.1,
                )
                parsed = _extract_json(result.get("content", ""))
                if isinstance(parsed, dict):
                    for dp in parsed.get("datapoints") or []:
                        all_evidence.append({**dp, "_source": "web_search"})

            # ── External CSV datasets ─────────────────────────────────────
            for ds_id in dataset_ids:
                rows = list(
                    self.db.scalars(
                        select(BenchmarkDatasetRow)
                        .where(BenchmarkDatasetRow.dataset_id == ds_id)
                        .limit(200)
                    )
                )
                if not rows:
                    continue
                row_summaries = [
                    {
                        "idx": i,
                        "title": r.normalized_title or "",
                        "level": r.normalized_level or "",
                        "loc": r.location or "",
                        "currency": r.currency or listing_ctx["currency"],
                        "base": r.base_salary,
                        "total": r.total_compensation,
                    }
                    for i, r in enumerate(rows)
                ]
                ext_system = (
                    "You are a compensation analyst reviewing an external salary dataset. "
                    "Identify the rows most similar to the job listing. "
                    "Return ONLY valid JSON:\n"
                    '{"matches": [{"idx": int, "confidence": float, "rationale": str}]}'
                )
                ext_result = await provider.generate(
                    system_prompt=ext_system,
                    messages=[{
                        "role": "user",
                        "content": (
                            f"Job listing:\n{json.dumps(listing_ctx, indent=2)}\n\n"
                            f"External dataset rows:\n{json.dumps(row_summaries, indent=2)}"
                        ),
                    }],
                    temperature=0.1,
                )
                ext_parsed = _extract_json(ext_result.get("content", ""))
                if isinstance(ext_parsed, dict):
                    for m in ext_parsed.get("matches") or []:
                        idx = m.get("idx")
                        if isinstance(idx, int) and 0 <= idx < len(rows):
                            r = rows[idx]
                            all_evidence.append({
                                "source_name": "External CSV",
                                "matched_title": r.normalized_title,
                                "geography": r.location,
                                "currency": r.currency or listing_ctx["currency"],
                                "p50_base": r.base_salary,
                                "total_comp_p50": r.total_compensation,
                                "confidence": float(m.get("confidence") or 0),
                                "citation_note": f"Dataset row {idx}",
                                "notes": str(m.get("rationale") or ""),
                                "_source": "external_csv",
                                "_dataset_id": str(ds_id),
                                "_row_idx": idx,
                            })

            # Persist evidence as matches
            confidence_scores: list[float] = []
            for ev in all_evidence:
                ds_id_val: uuid.UUID | None = None
                raw_ds = ev.get("_dataset_id")
                if raw_ds:
                    try:
                        ds_id_val = uuid.UUID(str(raw_ds))
                    except ValueError:
                        pass
                self.db.add(
                    BenchmarkMatch(
                        benchmark_run_id=run.id,
                        dataset_id=ds_id_val,
                        source_type=str(ev.get("_source", "web_search")),
                        matched_title=str(ev.get("matched_title") or ""),
                        matched_location=str(ev.get("geography") or ""),
                        base_salary=ev.get("p50_base"),
                        total_compensation=ev.get("total_comp_p50"),
                        currency=str(ev.get("currency") or listing_ctx["currency"]),
                        percentile="P50",
                        citation_url=None,
                        source_file_reference=str(ev.get("citation_note") or ""),
                        confidence_score=float(ev.get("confidence") or 0),
                        match_rationale=str(ev.get("notes") or ""),
                        raw_evidence_json={k: v for k, v in ev.items() if not k.startswith("_")},
                    )
                )
                confidence_scores.append(float(ev.get("confidence") or 0))

            p50_bases = [e.get("p50_base") for e in all_evidence if isinstance(e.get("p50_base"), (int, float))]
            summary: dict[str, Any] = {
                "sources_used": list(set(e.get("_source", "unknown") for e in all_evidence)),
                "total_datapoints": len(all_evidence),
                "market_p50_base": statistics.median(p50_bases) if p50_bases else None,
            }
            run.status = "completed"
            run.completed_at = datetime.now(timezone.utc)
            run.result_summary_json = summary
            run.confidence_score = statistics.mean(confidence_scores) if confidence_scores else 0.0
            self.db.flush()

        except Exception as exc:
            logger.exception("External benchmark run failed: %s", exc)
            run.status = "failed"
            run.result_summary_json = {"error": str(exc)}
            self.db.flush()

        return run

    # ── Recommendation chat ───────────────────────────────────────────────────

    async def chat_recommendation(
        self,
        job_listing_id: uuid.UUID,
        run_ids: list[uuid.UUID],
        messages: list[dict[str, str]],
        tenant_id: uuid.UUID,
    ) -> tuple[str, BenchmarkRecommendation | None]:
        listing = self.db.scalar(select(NegotiationCase).where(NegotiationCase.id == job_listing_id))
        if listing is None:
            raise ValueError("Job listing not found")

        listing_ctx = _listing_context(listing)

        # Build evidence context from completed runs
        evidence_context_parts: list[str] = []
        latest_run: JobListingBenchmarkRun | None = None
        for run_id in run_ids:
            run = self.db.scalar(
                select(JobListingBenchmarkRun)
                .where(JobListingBenchmarkRun.id == run_id)
                .where(JobListingBenchmarkRun.tenant_id == tenant_id)
            )
            if run and run.status == "completed":
                summary_str = json.dumps(run.result_summary_json or {}, indent=2)
                evidence_context_parts.append(
                    f"## {run.run_type.capitalize()} Benchmark Run ({run.id})\n{summary_str}"
                )
                latest_run = run

        evidence_context = "\n\n".join(evidence_context_parts) if evidence_context_parts else "No completed benchmark runs provided."

        system_prompt = (
            "You are a compensation advisor helping an HR operator set fair pay for a job listing. "
            "You have access to internal and external benchmark evidence. "
            "Help the user reason through options, risks, and tradeoffs.\n\n"
            f"## Job Listing\n{json.dumps(listing_ctx, indent=2)}\n\n"
            f"## Benchmark Evidence\n{evidence_context}\n\n"
            "When you are ready to make a final recommendation, include a JSON block in your response "
            "wrapped in <recommendation>...</recommendation> tags with this structure:\n"
            '{"recommended_base_min": float, "recommended_base_mid": float, "recommended_base_max": float, '
            '"recommended_total_comp_min": float|null, "recommended_total_comp_mid": float|null, '
            '"recommended_total_comp_max": float|null, "bonus_target": float|null, '
            '"equity_guidance": str|null, "currency": str, "location_basis": str, '
            '"confidence_score": float, "rationale": str, "caveats": str}\n\n'
            "Until you produce a final recommendation, keep the conversation grounded in the evidence. "
            "Be concise and actionable."
        )

        provider = get_provider()
        result = await provider.generate(
            system_prompt=system_prompt,
            messages=[{"role": m["role"], "content": m["content"]} for m in messages],
            temperature=0.2,
        )
        response_text: str = result.get("content", "")

        # Check if a recommendation was produced
        recommendation: BenchmarkRecommendation | None = None
        if "<recommendation>" in response_text and latest_run is not None:
            start = response_text.find("<recommendation>") + len("<recommendation>")
            end = response_text.find("</recommendation>")
            if end > start:
                rec_json_str = response_text[start:end].strip()
                rec_data = _extract_json(rec_json_str)
                if isinstance(rec_data, dict):
                    recommendation = BenchmarkRecommendation(
                        job_listing_id=job_listing_id,
                        benchmark_run_id=latest_run.id,
                        recommended_base_min=rec_data.get("recommended_base_min"),
                        recommended_base_mid=rec_data.get("recommended_base_mid"),
                        recommended_base_max=rec_data.get("recommended_base_max"),
                        recommended_total_comp_min=rec_data.get("recommended_total_comp_min"),
                        recommended_total_comp_mid=rec_data.get("recommended_total_comp_mid"),
                        recommended_total_comp_max=rec_data.get("recommended_total_comp_max"),
                        bonus_target=rec_data.get("bonus_target"),
                        equity_guidance=rec_data.get("equity_guidance"),
                        currency=rec_data.get("currency") or listing_ctx["currency"],
                        location_basis=rec_data.get("location_basis"),
                        confidence_score=rec_data.get("confidence_score"),
                        rationale=rec_data.get("rationale"),
                        caveats=rec_data.get("caveats"),
                        source_references_json=[str(r) for r in run_ids],
                    )
                    self.db.add(recommendation)
                    self.db.flush()

            # Strip the XML tags from user-visible response
            clean_text = response_text[:response_text.find("<recommendation>")].strip()
            if not clean_text:
                clean_text = response_text[response_text.find("</recommendation>") + len("</recommendation>"):].strip()
            response_text = clean_text or response_text

        return response_text, recommendation

    # ── Apply recommendation ─────────────────────────────────────────────────

    def apply_recommendation(
        self,
        recommendation_id: uuid.UUID,
        applied_by_id: uuid.UUID,
        tenant_id: uuid.UUID,
    ) -> NegotiationCase:
        rec = self.db.scalar(
            select(BenchmarkRecommendation)
            .join(JobListingBenchmarkRun, BenchmarkRecommendation.benchmark_run_id == JobListingBenchmarkRun.id)
            .where(BenchmarkRecommendation.id == recommendation_id)
            .where(JobListingBenchmarkRun.tenant_id == tenant_id)
        )
        if rec is None:
            raise ValueError("Recommendation not found")
        if rec.applied_to_listing:
            raise ValueError("Recommendation has already been applied")

        listing = self.db.scalar(
            select(NegotiationCase).where(NegotiationCase.id == rec.job_listing_id)
        )
        if listing is None:
            raise ValueError("Job listing not found")

        # Update the company confidential payload with the recommended compensation
        for party in listing.parties:
            if party.party_type == "company":
                confidential = dict(party.confidential_payload or {})
                if rec.recommended_base_min is not None:
                    confidential["budget_floor"] = rec.recommended_base_min
                if rec.recommended_base_max is not None:
                    confidential["budget_ceiling"] = rec.recommended_base_max
                if rec.recommended_base_mid is not None:
                    confidential["budget_target"] = rec.recommended_base_mid
                if rec.recommended_total_comp_max is not None:
                    confidential["total_comp_ceiling"] = rec.recommended_total_comp_max
                if rec.bonus_target is not None:
                    confidential["bonus_target_pct"] = rec.bonus_target
                confidential["benchmark_recommendation_id"] = str(recommendation_id)
                confidential["benchmark_applied_at"] = datetime.now(timezone.utc).isoformat()
                party.confidential_payload = confidential
                break

        rec.applied_to_listing = True
        rec.applied_at = datetime.now(timezone.utc)
        rec.applied_by = applied_by_id
        self.db.flush()
        return listing
